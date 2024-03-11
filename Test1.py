import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
class ScrollableFrame(tk.Frame):
    def __init__(self, master, **kwargs):
        tk.Frame.__init__(self, master, **kwargs)
        # Create canvas
        self.canvas = tk.Canvas(self, borderwidth=0, highlightthickness=0)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        # Add vertical scrollbar
        self.v_scrollbar = tk.Scrollbar(self, orient=tk.VERTICAL, command=self.canvas.yview)
        self.v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=self.v_scrollbar.set)
        # Create a frame inside the canvas which will be scrolled with it
        self.scrollable_frame = tk.Frame(self.canvas)
        self.canvas_frame_id = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor=tk.NW)
        # Bind events to update the scroll region and canvas width
        self.scrollable_frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)
    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
    def on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_frame_id, width=event.width)
class DataEntryForm:
    def __init__(self, root):
        self.root = root
        self.root.title("Data Entry Form")
        self.root.configure(bg="#f0f0f0")  # Set background color
        # Create Notebook widget
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        # Page 1: Data Entry Forms
        self.page1 = ScrollableFrame(self.notebook)
        self.notebook.add(self.page1, text="Data Entry Forms")
        self.create_forms(self.page1.scrollable_frame)
        # Page 2: Display Table
        self.page2 = ScrollableFrame(self.notebook)
        self.notebook.add(self.page2, text="Display Table")
        self.display_table(self.page2.scrollable_frame)
        # Page 3: Big Table
        self.page3 = ScrollableFrame(self.notebook)
        self.notebook.add(self.page3, text="Big Table")
        self.big_table(self.page3.scrollable_frame)

 
        submit_button = tk.Button(self.page3.scrollable_frame, text="Submit", command=self.submit_data, bg="#000080", fg="white", font=('Arial', 11, 'bold'))
        submit_button.pack(side="top", pady=20)  # Use pack instead of grid
    def create_forms(self, frame):
        # Create frame to contain all tables
        self.form_frame = tk.Frame(frame, bg="#f0f0f0")
        self.form_frame.pack(fill="both", expand=True)
        self.tables = [
            {'name': 'DCS Sizing Requirement', 'fields': ['Install Spare requirement(%)', 'Engineering Growth(%)', 'Rack Spares(%)', 'Cabinet Spare Space(%)', 'Serial Tag Packing']},
            {'name': 'DCS IO Summary', 'fields': ['Analog Inputs', 'Analog Outputs', 'Discrete Inputs', 'Discrete Outputs']},
            {'name': 'Additional Hardwired IO requirements', 'fields': ['Relay output Required', 'No. of Relay output required', 'Relay Output type']},
        ]
        self.tables_2 = [
            {'name': 'ESD IO Summary', 'fields': ['Install Spare requirement(%)', 'Engineering Growth(%)', 'Rack Spares(%)', 'Cabinet Spare Space(%)']},
            {'name': 'ESD IO Summary', 'fields': ['Analog Inputs', 'Analog Outputs', 'Discrete Inputs', 'Discrete Outputs']},
            {'name': 'Additional Hardwired IO requirements', 'fields': ['Relay output Required', 'No. of Relay output required', 'Relay Output type']},
        ]
        self.tables_3 = [
            {'name': 'ESD IO Summary', 'fields': ['Install Spare requirement(%)', 'Engineering Growth(%)', 'Rack Spares(%)', 'Cabinet Spare Space(%)']},
            {'name': 'ESD IO Summary', 'fields': ['Analog Inputs', 'Analog Outputs', 'Discrete Inputs', 'Discrete Outputs']},
            {'name': 'Additional Hardwired IO requirements', 'fields': ['Relay output Required', 'No. of Relay output required', 'Relay Output type']},
        ]
        self.entries = {}
        self.entries_2 = {}
        self.entries_3 = {}
        self.error_labels = {}
        self.error_labels_2 = {}
        self.error_labels_3 = {}
        for i, table in enumerate(self.tables):
            # Separator between tables
            if i > 0:
                separator = tk.Frame(self.form_frame, height=2, bd=1, relief=tk.SUNKEN, bg="#000080")
                separator.grid(row=6*i - 1, columnspan=3, sticky="ew", pady=(50, 0))  # Adjusted padding
            # Table label
            table_label = tk.Label(self.form_frame, text=table['name'], bg="#000080", fg="white", font=('Arial', 11, 'bold'))
            table_label.grid(row=6*i, columnspan=3, pady=(5, 2.5), sticky="ew")  # Span only 4 columns for the first table
            # Entries for the table
            self.entries[i] = []
            self.error_labels[i] = []
            for j, field in enumerate(table['fields']):
                label = tk.Label(self.form_frame, text=field, bg="#f0f0f0", font=('Arial', 9))
                label.grid(row=6*i + j + 1, column=0, padx=5, pady=2.5, sticky=tk.W)
                if i == 1:  # For table 2 (DCS IO Summary)
                    entry = tk.Entry(self.form_frame, validate="key")
                    entry.config(validatecommand=(self.root.register(self.validate_input), "%P"))
                    entry.grid(row=6*i + j + 1, column=1, padx=5, pady=2.5)
                    self.entries[i].append(entry)
                elif i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                    entry = tk.Entry(self.form_frame)
                    entry.grid(row=6*i + j + 1, column=1, padx=5, pady=2.5)
                    self.entries[i].append(entry)
                elif field == 'Serial Tag Packing':  # Dropdown for 'Serial Tag Packing'
                    options = ['10', '25', '50', '75', '100']
                    option_var = tk.StringVar(self.form_frame)
                    option_var.set(options[0])  # default value
                    option_menu = tk.OptionMenu(self.form_frame, option_var, *options)
                    option_menu.grid(row=6*i + j + 1, column=1, padx=5, pady=2.5)
                    self.entries[i].append(option_var)
                elif i == 2 and (field == 'Relay output Required' or field == 'Relay Output type'):  # Dropdowns for 'Relay output Required' and 'Relay Output type' in table 3
                    options = ['Yes', 'No'] if field == 'Relay output Required' else ['External', 'Integral']
                    option_var = tk.StringVar(self.form_frame)
                    option_var.set(options[0])  # default value
                    option_menu = tk.OptionMenu(self.form_frame, option_var, *options)
                    option_menu.grid(row=6*i + j + 1, column=1, padx=5, pady=2.5)
                    self.entries[i].append(option_var)
                else:
                    entry = tk.Entry(self.form_frame)
                    entry.grid(row=6*i + j + 1, column=1, padx=5, pady=2.5)
                    self.entries[i].append(entry)
                # Third column labels
                error_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", fg="red", font=('Arial', 9))
                error_label.grid(row=6*i + j + 1, column=2, padx=5, pady=2.5)
                self.error_labels[i].append(error_label)
                # Additional empty column
                empty_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", width=10)
                empty_label.grid(row=6*i + j + 1, column=3, padx=5, pady=2.5)
        for i, table in enumerate(self.tables_2):
                # Separator between tables
                if i > 0:
                    separator = tk.Frame(self.form_frame, height=2, bd=1, relief=tk.SUNKEN, bg="#bf8237")
                    separator.grid(row=6*i - 1, column=5, columnspan=3, sticky="ew", pady=(50, 0))  # Adjusted padding
                # Table label
                table_label = tk.Label(self.form_frame, text=table['name'], bg="#bf8237", fg="white", font=('Arial', 11, 'bold'))
                table_label.grid(row=6*i, column= 5, columnspan=3, pady=(5, 2.5), sticky="ew")  # Span only 4 columns for the first table
                # Entries for the table
                self.entries_2[i] = []
                self.error_labels_2[i] = []
                for j, field in enumerate(table['fields']):
                    label = tk.Label(self.form_frame, text=field, bg="#f0f0f0", font=('Arial', 9))
                    label.grid(row=6*i + j + 1, column=5, padx=5, pady=2.5, sticky=tk.W)
                    if i == 1:  # For table 2 (DCS IO Summary)
                        if j == 0:  # For the first textbox in the second row
                            entry = tk.Entry(self.form_frame, validate="key")
                            entry.config(validatecommand=(self.root.register(self.validate_input), "%P"))
                            entry.grid(row=6*i + j + 1, column=6, padx=5, pady=2.5)
                            self.entries_2[i].append(entry)
                        else:  # For the rest of the textboxes in the second row
                            entry = tk.Entry(self.form_frame)
                            entry.grid(row=6*i + j + 1, column=6, padx=5, pady=2.5)
                            self.entries_2[i].append(entry)
                    elif i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                        entry = tk.Entry(self.form_frame)
                        entry.grid(row=6*i + j + 1, column=6, padx=5, pady=2.5)
                        self.entries_2[i].append(entry)
                    elif i == 2 and (field == 'Relay output Required' or field == 'Relay Output type'):  # Dropdowns for 'Relay output Required' and 'Relay Output type' in table 3
                        options = ['Yes', 'No'] if field == 'Relay output Required' else ['External', 'Integral']
                        option_var = tk.StringVar(self.form_frame)
                        option_var.set(options[0])  # default value
                        option_menu = tk.OptionMenu(self.form_frame, option_var, *options)
                        option_menu.grid(row=6*i + j + 1, column=6, padx=5, pady=2.5)
                        self.entries_2[i].append(option_var)
                    else:
                        entry = tk.Entry(self.form_frame)
                        entry.grid(row=6*i + j + 1, column=6, padx=5, pady=2.5)
                        self.entries_2[i].append(entry)
                    # Third column labels
                    error_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", fg="red", font=('Arial', 9))
                    error_label.grid(row=6*i + j + 1, column=7, padx=5, pady=2.5)
                    self.error_labels_2[i].append(error_label)
                    # Additional empty column
                    empty_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", width=10)
                    empty_label.grid(row=6*i + j + 1, column=8, padx=5, pady=2.5)
        for i, table in enumerate(self.tables_3):
                # Separator between tables
                if i > 0:
                    separator = tk.Frame(self.form_frame, height=2, bd=1, relief=tk.SUNKEN, bg="#2fa347")
                    separator.grid(row=6*i - 1, column=9, columnspan=3, sticky="ew", pady=(50, 0))  # Adjusted padding
                # Table label
                table_label = tk.Label(self.form_frame, text=table['name'], bg="#2fa347", fg="white", font=('Arial', 11, 'bold'))
                table_label.grid(row=6*i, column= 9, columnspan=3, pady=(5, 2.5), sticky="ew")  # Span only 4 columns for the first table
                # Entries for the table
                self.entries_3[i] = []
                self.error_labels_3[i] = []
                for j, field in enumerate(table['fields']):
                    label = tk.Label(self.form_frame, text=field, bg="#f0f0f0", font=('Arial', 9))
                    label.grid(row=6*i + j + 1, column=9, padx=5, pady=2.5, sticky=tk.W)
                    if i == 1:  # For table 2 (DCS IO Summary)
                        if j == 0:  # For the first textbox in the second row
                            entry = tk.Entry(self.form_frame, validate="key")
                            entry.config(validatecommand=(self.root.register(self.validate_input), "%P"))
                            entry.grid(row=6*i + j + 1, column=10, padx=5, pady=2.5)
                            self.entries_3[i].append(entry)
                        else:  # For the rest of the textboxes in the second row
                            entry = tk.Entry(self.form_frame)
                            entry.grid(row=6*i + j + 1, column=10, padx=5, pady=2.5)
                            self.entries_3[i].append(entry)
                    elif i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                        entry = tk.Entry(self.form_frame)
                        entry.grid(row=6*i + j + 1, column=10, padx=5, pady=2.5)
                        self.entries_3[i].append(entry)
                    elif i == 2 and (field == 'Relay output Required' or field == 'Relay Output type'):  # Dropdowns for 'Relay output Required' and 'Relay Output type' in table 3
                        options = ['Yes', 'No'] if field == 'Relay output Required' else ['External', 'Integral']
                        option_var = tk.StringVar(self.form_frame)
                        option_var.set(options[0])  # default value
                        option_menu = tk.OptionMenu(self.form_frame, option_var, *options)
                        option_menu.grid(row=6*i + j + 1, column=10, padx=5, pady=2.5)
                        self.entries_3[i].append(option_var)
                    else:
                        entry = tk.Entry(self.form_frame)
                        entry.grid(row=6*i + j + 1, column=10, padx=5, pady=2.5)
                        self.entries_3[i].append(entry)
                    # Third column labels
                    error_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", fg="red", font=('Arial', 9))
                    error_label.grid(row=6*i + j + 1, column=11, padx=5, pady=2.5)
                    self.error_labels_3[i].append(error_label)
 

    def display_table(self, frame):
        # Create frame to contain all tables
        self.form_frame = tk.Frame(frame, bg="#f0f0f0")
        self.form_frame.pack(fill="both", expand=True)
        self.tables4 = [
            {'name': 'Modbus TCP/IP Interface', 'fields': ['Serial IO tag Count', 'Serial Link Count', 'Redundancy Required']},
            {'name': 'Ethernet TCP/IP Interface', 'fields': ['Serial IO tag Count', 'Serial Link Count', 'Redundancy Required']},
            {'name': 'IEC 61850 Interface', 'fields': ['Serial IO tag Count', 'Serial Link Count', 'Redundancy Required']},
            {'name': 'OPC UA Interface', 'fields': ['Serial IO tag Count', 'Serial Link Count', 'Redundancy Required']},
            {'name': 'Ethernet TCP/IP Control Tag Integration', 'fields': ['Serial IO tag Count', 'Serial Link Count', 'Redundancy Required']},
        ]
        self.tables6 = [
            {'name': 'Third Party OPC Integration (to OPC Server)', 'fields': ['OPC tag Count (Additional)', 'Third Party OPC Packages']},
            {'name': 'Controller Sizing', 'fields': ['Type of Controllers', 'Simplex or Redundant']},
        ]
        self.entries4 = {}
        self.error_labels4 = {}
        self.entries6 = {}
        self.error_labels6 = {}
        for i, table in enumerate(self.tables4):
            # Separator between tables
            if i > 0:
                separator = tk.Frame(self.form_frame, height=2, bd=1, relief=tk.SUNKEN, bg="#000080")
                separator.grid(row=6*i - 1, columnspan=3, sticky="ew", pady=(50, 0))  # Adjusted padding
            # Table label
            table_label = tk.Label(self.form_frame, text=table['name'], bg="#000080", fg="white", font=('Arial', 11, 'bold'))
            table_label.grid(row=6*i, columnspan=3, pady=(5, 2.5), sticky="ew")  # Span only 4 columns for the first table
            # Entries for the table
            self.entries4[i] = []
            self.error_labels4[i] = []
            for j, field in enumerate(table['fields']):
                label = tk.Label(self.form_frame, text=field, bg="#f0f0f0", font=('Arial', 9))
                label.grid(row=6*i + j + 1, column=0, padx=5, pady=2.5, sticky=tk.W)
                if field == 'Redundancy Required':  # Dropdown for 'Redundancy Required'
                    redundancy_options = ['Yes', 'No']
                    redundancy_var = tk.StringVar(self.form_frame)
                    redundancy_var.set(redundancy_options[0])  # default value
                    redundancy_menu = tk.OptionMenu(self.form_frame, redundancy_var, *redundancy_options)
                    redundancy_menu.grid(row=6 * i + j + 1, column=1, padx=5, pady=2.5)
                    self.entries4[i].append(redundancy_var)
                elif field == 'Third Party OPC Packages':  # Dropdown for 'Third Party OPC Packages'
                    opc_options = [str(k) for k in range(1, 51)]  # Numbers 1-50
                    opc_var = tk.StringVar(self.form_frame)
                    opc_var.set(opc_options[0])  # default value
                    opc_menu = tk.OptionMenu(self.form_frame, opc_var, *opc_options)
                    opc_menu.grid(row=6 * i + j + 1, column=1, padx=5, pady=2.5)
                    self.entries4[i].append(opc_var)
                else:
                    entry = tk.Entry(self.form_frame)
                    entry.grid(row=6*i + j + 1, column=1, padx=5, pady=2.5)
                    self.entries4[i].append(entry)
                # Third column labels
                error_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", fg="red", font=('Arial', 9))
                error_label.grid(row=6*i + j + 1, column=2, padx=5, pady=2.5)
                self.error_labels4[i].append(error_label)
                # Additional empty column
                empty_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", width=10)
                empty_label.grid(row=6*i + j + 1, column=3, padx=5, pady=2.5)
            for i, table in enumerate(self.tables6):
            # Separator between tables
                if i > 0:
                    separator = tk.Frame(self.form_frame, height=2, bd=1, relief=tk.SUNKEN, bg="#000080")
                    separator.grid(row=6*i - 1, columnspan=3, column=5, sticky="ew", pady=(50, 0))  # Adjusted padding
                # Table label
                table_label = tk.Label(self.form_frame, text=table['name'], bg="#000080", fg="white", font=('Arial', 11, 'bold'))
                table_label.grid(row=6*i, columnspan=3, column= 5,pady=(5, 2.5), sticky="ew")  # Span only 4 columns for the first table
                # Entries for the table
                self.entries6[i] = []
                self.error_labels6[i] = []
                for j, field in enumerate(table['fields']):
                    label = tk.Label(self.form_frame, text=field, bg="#f0f0f0", font=('Arial', 9))
                    label.grid(row=6*i + j + 1, column=5, padx=5, pady=2.5, sticky=tk.W)
                    if field == 'Third Party OPC Packages':  # Dropdown for 'Third Party OPC Packages'
                        opc_options = [str(k) for k in range(1, 51)]  # Numbers 1-50
                        opc_var = tk.StringVar(self.form_frame)
                        opc_var.set(opc_options[0])  # default value
                        opc_menu = tk.OptionMenu(self.form_frame, opc_var, *opc_options)
                        opc_menu.grid(row=6 * i + j + 1, column=6, padx=5, pady=2.5)
                        self.entries6[i].append(opc_var)
                    elif field == 'Type of Controllers':  # Dropdown for 'Redundancy Required'
                        c_options = ['MQ controller', 'MX controller', 'SX controller', 'SQ controller']
                        c_var = tk.StringVar(self.form_frame)
                        c_var.set(c_options[0])  # default value
                        c_menu = tk.OptionMenu(self.form_frame, c_var, *c_options)
                        c_menu.grid(row=6 * i + j + 1, column=6, padx=5, pady=2.5)
                        self.entries6[i].append(c_var)
                    elif field == 'Simplex or Redundant':  # Dropdown for 'Redundancy Required'
                        s_options = ['Redundant', 'Simplex']
                        s_var = tk.StringVar(self.form_frame)
                        s_var.set(s_options[0])  # default value
                        s_menu = tk.OptionMenu(self.form_frame, s_var, *s_options)
                        s_menu.grid(row=6 * i + j + 1, column=6, padx=5, pady=2.5)
                        self.entries6[i].append(s_var)
                    else:
                        entry = tk.Entry(self.form_frame)
                        entry.grid(row=6*i + j + 1, column=6, padx=5, pady=2.5)
                        self.entries6[i].append(entry)
                    # Third column labels
                    error_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", fg="red", font=('Arial', 9))
                    error_label.grid(row=6*i + j + 1, column=7, padx=5, pady=2.5)
                    self.error_labels6[i].append(error_label)

    def big_table(self, frame):
        # Create frame to contain all tables
        self.form_frame = tk.Frame(frame, bg="#f0f0f0")
        self.form_frame.pack(fill="both", expand=True)
        self.tables5 = [
            {'name': 'Server / Workstations', 'fields': ['Proplus Server', 'Engineering Server', 'Historian Server', 'Asset Management System', 'OPC Server', 'Operator Station', 'Alarm Mgmt System (AgileOps)', 'Domain Controller', 'Backup & Recovery System', 'ePO Server', 'Large Screen Display', 'Black & White Printer', 'Color Printer', 'GPS Server including Antenna']},
        ]
        self.error_labels5 = {}  # Add this line
        self.entries5 = {}

 
        for i, table in enumerate(self.tables5):
            # Separator between tables
            if i > 0:
                separator = tk.Frame(self.form_frame, height=2, bd=1, relief=tk.SUNKEN, bg="#000080")
                separator.grid(row=6*i - 1, columnspan=7, sticky="ew", pady=(50, 0))  # Adjusted padding
 
            # Table labels
            table_label = tk.Label(self.form_frame, text=table['name'], bg="#000080", fg="white", font=('Arial', 11, 'bold'))
            table_label.grid(row=6*i, column=0, padx=5, pady=(5, 2.5), sticky="ew")  # Server n Workstation column
 
            labels = ['Required', 'Quantities', 'M/C Type', 'Monitor Required', 'Qty of Monitors', 'KVM Required']
            for j, label_text in enumerate(labels):
                label = tk.Label(self.form_frame, text=label_text, bg="#000080", fg="white", font=('Arial', 11, 'bold'))
                label.grid(row=6*i, column=j+1, columnspan=1, padx=5, pady=(5, 2.5), sticky="ew")
 
            # Entries for the table
            self.entries5[i] = []
            self.error_labels5[i] = []
 
            for j, field in enumerate(table['fields']):
                label = tk.Label(self.form_frame, text=field, bg="#f0f0f0", font=('Arial', 9))
                label.grid(row=6*i + j + 1, column=0, padx=5, pady=2.5, sticky=tk.W)
 
                for k, label_text in enumerate(labels):
                    if label_text in ['Required', 'Monitor Required']:
                        option_var = tk.StringVar(self.form_frame)
                        options = ['Yes', 'No']
                        option_menu = tk.OptionMenu(self.form_frame, option_var, *options)
                        option_menu.grid(row=6*i + j + 1, column=k+1, padx=5, pady=2.5)
                        option_var.set(options[0])  # Set default value
                        self.entries5[i].append(option_var)
                    elif label_text == 'M/C Type':
                        option_var = tk.StringVar(self.form_frame)
                        options = ['----', 'Tower', 'Rack']
                        option_menu = tk.OptionMenu(self.form_frame, option_var, *options)
                        option_menu.grid(row=6*i + j + 1, column=k+1, padx=5, pady=2.5)
                        option_var.set(options[0])  # Set default value
                        self.entries5[i].append(option_var)
                    elif label_text == 'KVM Required':
                        value = '1' if j in [1, 3, 8] else '0' if j in [0, 2, 4, 5, 6, 7, 9] else '-'  # Assigning values based on row index
                        label = tk.Label(self.form_frame, text=value, bg="#f0f0f0", font=('Arial', 9))
                        label.grid(row=6*i + j + 1, column=k+1, padx=5, pady=2.5)
                        self.entries5[i].append(label)
 
                    else:
                        entry = tk.Entry(self.form_frame)
                        entry.grid(row=6*i + j + 1, column=k+1, padx=5, pady=2.5)
                        self.entries5[i].append(entry)
 
 
 

            # Third column labels
            error_label = tk.Label(self.form_frame, text="", bg="#f0f0f0", fg="red", font=('Arial', 9))
            error_label.grid(row=6*i + j + 1, column=7, padx=5, pady=2.5)
            self.error_labels5[i].append(error_label)   

    def validate_input(self, new_text):
        if new_text == "":
            return True  # Allow backspacing
        return len(new_text) <= 6 and new_text.isdigit()
    def submit_data(self):
        error_flag = False
        for i, table in enumerate(self.tables):
            for j, field in enumerate(table['fields']):
                entry = self.entries[i][j].get().strip()  # Remove leading/trailing spaces
                if i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                    if not entry.isdigit():
                        self.display_error(i, j)
                        error_flag = True
                    else:
                        self.error_labels[i][j].config(text="")  # Reset error message for the first table
                elif not entry.isdigit() and field not in ['Serial Tag Packing', 'Relay output Required', 'Relay Output type', 'Redundancy Required']:
                    self.display_error(i, j)
                    error_flag = True
                else:
                    self.error_labels[i][j].config(text="")  # Reset error message for other tables
        for i, table in enumerate(self.tables_2):
            for j, field in enumerate(table['fields']):
                entry = self.entries_2[i][j].get().strip()  # Remove leading/trailing spaces
                if i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                    if not entry.isdigit():
                        self.display_error2(i, j)
                        error_flag = True
                    else:
                        self.error_labels_2[i][j].config(text="")  # Reset error message for the first table
                elif not entry.isdigit() and field not in ['Relay output Required', 'Relay Output type']:
                    self.display_error2(i, j)
                    error_flag = True
                else:
                    self.error_labels_2[i][j].config(text="")  # Reset error message for other tables
        for i, table in enumerate(self.tables_3):
            for j, field in enumerate(table['fields']):
                entry = self.entries_3[i][j].get().strip()  # Remove leading/trailing spaces
                if i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                    if not entry.isdigit():
                        self.display_error3(i, j)
                        error_flag = True
                    else:
                        self.error_labels_3[i][j].config(text="")  # Reset error message for the first table
                elif not entry.isdigit() and field not in ['Relay output Required', 'Relay Output type']:
                    self.display_error3(i, j)
                    error_flag = True
                else:
                    self.error_labels_3[i][j].config(text="")  # Reset error message for other tables
        for i, table in enumerate(self.tables4):
            for j, field in enumerate(table['fields']):
                entry = self.entries4[i][j].get().strip()  # Remove leading/trailing spaces
                if entry == '' and field in ['Serial IO tag Count', 'Serial Link Count']:
                    # Check if both "Serial IO tag Count" and "Serial Link Count" are empty, allow submission
                    if self.entries4[i][0].get().strip() == '' and self.entries4[i][1].get().strip() == '':
                        continue
                    else:
                        self.display_empty_error(i, j)
                        error_flag = True
                elif i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                    if not entry.isdigit():
                        self.display_error4(i, j)
                        error_flag = True
                    else:
                        self.error_labels4[i][j].config(text="")  # Reset error message for the first table
                elif not entry.isdigit() and field not in ['Redundancy Required']:
                    self.display_error4(i, j)
                    error_flag = True
                else:
                    self.error_labels4[i][j].config(text="")  # Reset error message for other tables
        for i, table in enumerate(self.tables6):
            for j, field in enumerate(table['fields']):
                entry = self.entries6[i][j].get().strip()  # Remove leading/trailing spaces
                if i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                    if not entry.isdigit():
                        self.display_error6(i, j)
                        error_flag = True
                    else:
                        self.error_labels6[i][j].config(text="")  # Reset error message for the first table
                elif not entry.isdigit() and field not in ['Type of Controllers','Simplex or Redundant','Third Party OPC Packages']:
                    self.display_error6(i, j)
                    error_flag = True
                else:
                    self.error_labels6[i][j].config(text="")   # Reset error message for other tables
        for i, table in enumerate(self.tables5):
            for j, field in enumerate(table['fields']):
                entry = self.entries5[i][j].get().strip()  # Remove leading/trailing spaces
                if i == 0 and j != len(table['fields']) - 1:  # For the first table, except the last field
                    if not entry.isdigit():
                        self.display_error5(i, j)
                        error_flag = True
                    else:
                        self.error_labels5[i][j].config(text="")  # Reset error message for the first table
                elif not entry.isdigit() and field not in ['Required']:
                    self.display_error5(i, j)
                    error_flag = True
                else:
                    self.error_labels5[i][j].config(text="")  # Reset error message for other tables

        if not error_flag:
            all_data = []
            for i, table in enumerate(self.tables):
                table_data = []
                for j, field in enumerate(table['fields']):
                    entry = self.entries[i][j].get()
                    table_data.append(entry)
                all_data.append(table_data)
            for i, table in enumerate(self.tables_2):
                table_data = []
                for j, field in enumerate(table['fields']):
                    entry = self.entries_2[i][j].get()
                    table_data.append(entry)
                all_data.append(table_data)
            for i, table in enumerate(self.tables_3):
                table_data = []
                for j, field in enumerate(table['fields']):
                    entry = self.entries_3[i][j].get()
                    table_data.append(entry)
                all_data.append(table_data)
            for i, table in enumerate(self.tables4):
                table_data = []
                for j, field in enumerate(table['fields']):
                    entry = self.entries4[i][j].get()
                    table_data.append(entry)
                all_data.append(table_data)
            for i, table in enumerate(self.tables6):
                table_data = []
                for j, field in enumerate(table['fields']):
                    entry = self.entries6[i][j].get()
                    table_data.append(entry)
                all_data.append(table_data)
            for i, table in enumerate(self.tables5):
                table_data = []
                for j, field in enumerate(table['fields']):
                    entry = self.entries5[i][j].get()
                    table_data.append(entry)
                all_data.append(table_data)
            if all_data:
                self.write_to_excel(all_data)
                messagebox.showinfo("Success", "Data saved to Excel successfully!")
            else:
                messagebox.showerror("Error", "Please fill in all fields.")
    def display_empty_error(self, table_index, row_index):
        if row_index == 0:
            self.error_labels4[table_index][row_index].config(text="*Tag count is missing")
        elif row_index == 1:
            self.error_labels4[table_index][row_index].config(text="*Third party link count is missing")
    def write_to_excel(self, data):
        try:
            wb = openpyxl.load_workbook('Tables.xlsx')
            sheet = wb.active
            sheet.title = "DataSheet"
            # Write data to specific cells based on the provided logic
            sheet.cell(row=5, column=3).value = int(data[0][0]) / 100  # Install Spare requirement
            sheet.cell(row=6, column=3).value = int(data[0][1]) / 100  # Engineering Growth
            sheet.cell(row=7, column=3).value = int(data[0][2]) / 100  # Rack Spares
            sheet.cell(row=8, column=3).value = int(data[0][3]) / 100  # Cabinet Spare Space
            sheet.cell(row=9, column=3).value = data[0][4]  # Serial Tag Packing
            sheet.cell(row=13, column=3).value = data[1][0]  # Analog Inputs
            sheet.cell(row=14, column=3).value = data[1][1]  # Analog Outputs
            sheet.cell(row=15, column=3).value = data[1][2]  # Discrete Inputs
            sheet.cell(row=16, column=3).value = data[1][3]  # Discrete Outputs
            sheet.cell(row=19, column=3).value = data[2][0]  # Relay output Required
            sheet.cell(row=20, column=3).value = data[2][1]  # Number of relay output required
            sheet.cell(row=21, column=3).value = data[2][2]  # Relay Output type
 

            # Data from page 2
            sheet.cell(row=24, column=3).value = data[3][0]  # Serial IO tag Count (Page 2)
            sheet.cell(row=25, column=3).value = data[3][1]  # Serial Link Count (Page 2)
            sheet.cell(row=26, column=3).value = data[3][2]  # Redundancy Required (Page 2)
            wb.save("Tables.xlsx")
            messagebox.showinfo("Success", "Data saved to Excel successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to write data to Excel: {e}")
    def display_error(self, table_index, row_index):
        self.error_labels[table_index][row_index].config(text="*Please enter an integer")
    def display_error2(self, table_index, row_index):
        self.error_labels_2[table_index][row_index].config(text="*Please enter an integer2")
    def display_error3(self, table_index, row_index):
        self.error_labels_3[table_index][row_index].config(text="*Please enter an integer3")
    def display_error4(self, table_index, row_index):
        self.error_labels4[table_index][row_index].config(text="*Please enter an integer4")
    def display_error5(self, table_index, row_index):
        self.error_labels5[table_index][row_index].config(text="*Please enter an integer5")
    def display_error6(self, table_index, row_index):
        self.error_labels6[table_index][row_index].config(text="*Please enter an integer6")
 

if __name__ == "__main__":
    root = tk.Tk()
    app = DataEntryForm(root)
    root.mainloop()